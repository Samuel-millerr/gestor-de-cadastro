import tkinter as tk
from openpyxl import Workbook, load_workbook
from tkinter import messagebox

def mostrar_inter_cliente():

    def inserir_cliente():
        nome_cliente = case1.get()
        data_nasc = case2.get()
        cpf = case3.get()
        origem = case4.get()
        score = case5.get()
        cliente.append({
            'Nome' : nome_cliente, 
            'Data de Nascimento' : data_nasc,
            'CPF' : cpf, 
            'Origem' : origem, 
            'Score' : score
        })
        
        case1.delete(0, tk.END)
        case2.delete(0, tk.END)
        case3.delete(0, tk.END)
        case4.delete(0, tk.END)
        case5.delete(0, tk.END)

        
    def salvar_em_excel(): 
        try:
            workbook = load_workbook('clientes.xlsx')
            sheet = workbook.active 
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active 
            sheet.append(['Nome', 'Data de Nascimento', 'CPF', 'Origem', 'Score'])
   
        for client in cliente:
            sheet.append([client['Nome'], client['Data de Nascimento'], client['CPF'], client['Origem'], client['Score']])
    
        workbook.save("clientes.xlsx")
        
        messagebox.showinfo("Sucesso", "Dados salvos no excel!")
    
    
    def consultar_dados():
        try:
            workbook = load_workbook('clientes.xlsx')
            sheet = workbook.active
            dados = []
            for row in sheet.iter_rows(min_row = 2, values_only = True):
                nomes = casex.get()
                if nomes == row[0]:
                    dados.append(f"Nome: {row[0]}, Data de Nascimento: {row[1]}, CPF: {row[2]}, Origem: {row[3]}, Score: {row[4]}")
            messagebox.showinfo("Dados dos Clientes", "\n".join(dados))
        except FileNotFoundError:
            messagebox.showerror("Erro", "Arquivo 'clientes.xlsx' não encontrado.")
                    
                    
    window = tk.Tk()
    window.title("Cadastro Cliente")

    cliente = []
    
    l1 = tk.Label(window, text = "Digite as informações do cliente a seguir:")
    l1.grid(row = 0, column= 0,pady= 2)

    l2 = tk.Label(window, text = "Nome:")
    l2.grid(row = 1, column=0, pady = 2)
    case1 = tk.Entry(window)
    case1.grid(row = 1, column= 1, pady=2)

    l3 = tk.Label(window, text = "Data de Nascimento:")
    l3.grid(row = 2, column=0, pady = 2)
    case2 = tk.Entry(window)
    case2.grid(row=2, column=1, pady = 2)

    l4 = tk.Label(window, text = "CPF:")
    l4.grid(row = 3, column = 0, pady = 2)
    case3 = tk.Entry(window)
    case3.grid(row = 3, column = 1, pady = 2)

    l5 = tk.Label(window, text = "Origem:")
    l5.grid(row = 4, column = 0, pady = 1)
    case4 = tk.Entry(window)
    case4.grid(row = 4, column = 1, pady = 2)

    l6 = tk.Label(window, text = "Score:")
    l6.grid(row = 5, column = 0, pady = 2)
    case5 = tk.Entry(window)
    case5.grid(row = 5, column = 1, pady = 2)
           
    button = tk.Button(window, text = "Inserir", command = inserir_cliente)
    button.grid(row = 6, column = 1, pady = 2)
    
    button = tk.Button(window, text = "Adicionar ao Exel", command = salvar_em_excel)
    button.grid(row = 7, column = 1, pady =2)
    
    casex = tk.Entry(window)
    casex.grid(row = 8, column = 1, pady = 2)
    
    button = tk.Button(window, text = "Consultar Cliente", command = consultar_dados)
    button.grid(row = 9, column = 1, pady =2)

    window.geometry("400x265")
    window.mainloop()
    
    

def mostrar_interface_produto():

    def inserir_produto():
        nomeProd = case1.get()
        valor = case2.get()
        categ = case3.get()
        prod.append({'Nome' : nomeProd,
                     'Valor' : valor,
                     'Categoria' : categ
        })
        
        case1.delete(0, tk.END)
        case2.delete(0, tk.END)
        case3.delete(0, tk.END)

    
    def salvar_em_excel(): 
        try:
            workbook = load_workbook('produtos.xlsx')
            sheet = workbook.active 
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(['Nome', 'Valor', 'Categoria'])
   
        for produto in prod:
            sheet.append([produto['Nome'], produto['Valor'], produto['Categoria']])
        workbook.save("produtos.xlsx")
            
        messagebox.showinfo("Sucesso","Dados salvos no excel!")
        
    def consultar_dados():
        try:
            workbook = load_workbook('produtos.xlsx')
            sheet = workbook.active
            dados = []
            for row in sheet.iter_rows(min_row = 2, values_only = True):
                nomes = casex.get()
                if nomes == row[0]:
                    dados.append(f"Nome: {row[0]}, Valor: {row[1]}, Categoria: {row[2]}")
            messagebox.showinfo("Dados dos Produtos", "\n".join(dados))
        except FileNotFoundError:
            messagebox.showerror("Erro", "Arquivo 'produtos.xlsx' não encontrado.")
            
        
    window = tk.Tk()
    window.title("Cadastro Produto")

    l1 = tk.Label(window, text = "Digite as informações do produto: ")
    l1.grid(row = 0, column = 0, pady = 2)

    l2 = tk.Label(window, text = "Nome do Produto: ")
    l2.grid(row = 1, column = 0, pady = 2)
    case1 = tk.Entry(window)
    case1.grid(row = 1, column = 1, pady =2)

    l3 = tk.Label(window, text = "Valor:")
    l3.grid(row = 2, column = 0, pady = 2)
    case2 = tk.Entry(window)
    case2.grid(row = 2, column = 1, pady = 2)

    l4 = tk.Label(window, text = "Categoria:")
    l4.grid(row = 3, column = 0, pady = 2)
    case3 = tk.Entry(window)
    case3.grid(row = 3, column = 1, pady = 2)

    button = tk.Button(window, text = "Inserir", command = inserir_produto)
    button.grid(row = 4, column = 1, pady = 2)
    
    button = tk.Button(window, text = "Adicionar ao Exel", command = salvar_em_excel)
    button.grid(row = 5, column = 1, pady = 2)

    casex = tk.Entry(window)
    casex.grid(row = 6, column = 1, pady = 2)
    
    button = tk.Button(window, text = "Consultar", command = consultar_dados)
    button.grid(row = 7, column = 1,pady = 2)
    
    prod = []
    
    window.geometry("340x215")
    window.mainloop()

def mostar_interface_vendedor():
    vend = []
    
    def inserir_vendedor():
        nome = case1.get()
        matri = case2.get()
        vend.append({'Nome' : nome,
                     'Matrícula' : matri
        })
        
        case1.delete(0, tk.END)
        case2.delete(0, tk.END)  
       
    def salvar_no_excel(): 
        try:
            workbook = load_workbook('vendedor.xlsx')
            sheet = workbook.active 
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(['Nome', 'Matrícula'])
   
        for vendedor in vend:
            sheet.append([vendedor['Nome'], vendedor['Matrícula']])
        workbook.save("vendedor.xlsx")
            
        messagebox.showinfo("Sucesso","Dados salvos no exel!")
            
    def consultar_dados():
        try:
            workbook = load_workbook('vendedor.xlsx')
            sheet = workbook.active
            dados = []
            for row in sheet.iter_rows(min_row = 2, values_only = True):
                nomes = casex.get()
                if nomes == row[0]:
                    dados.append(f"Nome: {row[0]}, Matrícula: {row[1]}")
            messagebox.showinfo("Dados dos Vendedores", "\n".join(dados))
        except FileNotFoundError:
            messagebox.showerror("Erro", "Arquivo 'vendedor.xlsx' não encontrado.")
            
    window = tk.Tk()
    window.title("Cadastro Vendedor")

    l1 = tk.Label(window, text = "Insira os dados do vendedor: ")
    l1.grid(row = 0, column = 0, pady = 2)

    l2 = tk.Label(window, text = "Nome do Vendedor:")
    l2.grid(row = 1, column = 0, pady = 2)
    case1 = tk.Entry(window)
    case1.grid(row = 1, column = 1, pady = 2)

    l3 = tk.Label(window, text = "Matrícula:")
    l3.grid(row = 2, column = 0, pady = 2)
    case2 = tk.Entry(window)
    case2.grid(row = 2, column = 1, pady = 2)

    button = tk.Button(window, text = "Inserir", command = inserir_vendedor)
    button.grid(row = 3, column = 1, pady = 2)

    button = tk.Button(window, text = "Adicionar ao Exel", command = salvar_no_excel)
    button.grid(row = 4, column = 1, pady = 2)
    
    casex = tk.Entry(window)
    casex.grid(row = 5, column = 1, pady = 2)
    
    button = tk.Button(window, text = "Consultar", command = consultar_dados)
    button.grid(row = 6, column = 1, pady = 2)
    
    
    window.geometry("300x195")
    window.mainloop()

window = tk.Tk()
window.title("Cadastro")

l1 = tk.Label(window, text = "Essas são as três opções para inserir informações")
l1.grid(row = 0, column = 0, pady = 2)
button1 = tk.Button(window, text = "Cliente", command = mostrar_inter_cliente)
button1.grid(row = 1, column = 0, pady = 2)
button2 = tk.Button(window, text = "Produto", command = mostrar_interface_produto)
button2.grid(row = 2, column = 0, pady = 2)
button3 = tk.Button(window, text = "Vendedor",command = mostar_interface_vendedor)
button3.grid(row = 3, column = 0, pady = 2)

window.geometry("262x135")
window.mainloop()
